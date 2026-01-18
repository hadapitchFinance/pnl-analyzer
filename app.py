import calendar
import csv
import hashlib
import io
import json
import os
import re
import sqlite3
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import streamlit as st

st.set_page_config(page_title="Trade Analyzer (IBKR + Fidelity)", layout="wide")

st.markdown(
    """
    <style>
        :root {
            --bg-card: rgba(15, 23, 42, 0.5);
            --border-subtle: rgba(148, 163, 184, 0.12);
            --text-muted: #94a3b8;
        }
        .section-title {
            font-size: 1.25rem;
            font-weight: 700;
            margin: 1.2rem 0 0.6rem 0;
        }
        .dashboard-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(190px, 1fr));
            gap: 0.9rem;
        }
        .stat-card {
            background: linear-gradient(145deg, #101827, #0d1522);
            border: 1px solid rgba(148, 163, 184, 0.18);
            border-radius: 16px;
            padding: 16px 18px;
            box-shadow: 0 10px 24px rgba(15, 23, 42, 0.28);
        }
        .stat-card .stat-label {
            color: #93c5fd;
            font-size: 0.85rem;
            letter-spacing: 0.02em;
            font-weight: 600;
            text-transform: uppercase;
            margin-bottom: 6px;
        }
        .stat-card .stat-value {
            color: #f8fafc;
            font-size: 1.55rem;
            font-weight: 700;
        }
        .stat-card .stat-sub {
            color: #cbd5f5;
            font-size: 0.85rem;
            margin-top: 6px;
        }
        .stat-sub {
            color: #cbd5f5;
            font-size: 0.8rem;
            margin-top: 6px;
        }
        .section-card {
            background: var(--bg-card);
            border: 1px solid var(--border-subtle);
            border-radius: 18px;
            padding: 16px 18px 6px;
            margin-bottom: 16px;
        }
        .status-card {
            background: rgba(15, 23, 42, 0.75);
            border: 1px solid rgba(148, 163, 184, 0.18);
            border-radius: 16px;
            padding: 14px 16px;
        }
        .status-label {
            font-size: 0.78rem;
            letter-spacing: 0.03em;
            text-transform: uppercase;
            color: var(--text-muted);
            margin-bottom: 6px;
        }
        .status-value {
            font-size: 1.1rem;
            font-weight: 700;
            color: #e2e8f0;
        }
        .status-clean {
            color: #22c55e;
        }
        .status-warn {
            color: #f59e0b;
        }
        .pill {
            display: inline-block;
            padding: 4px 10px;
            border-radius: 999px;
            font-size: 0.75rem;
            font-weight: 600;
            background: rgba(15, 23, 42, 0.6);
            border: 1px solid rgba(148, 163, 184, 0.2);
        }
    </style>
    """,
    unsafe_allow_html=True,
)

def render_stat_cards(items: List[Dict[str, str]]) -> None:
    cards = "".join(
        [
            "<div class='stat-card'>"
            f"<div class='stat-label'>{item['label']}</div>"
            f"<div class='stat-value'>{item['value']}</div>"
            f"<div class='stat-sub'>{item.get('sub', '')}</div>"
            "</div>"
            for item in items
        ]
    )
    st.markdown(f"<div class='dashboard-grid'>{cards}</div>", unsafe_allow_html=True)

def render_status_cards(items: List[Dict[str, str]]) -> None:
    cards = "".join(
        [
            "<div class='status-card'>"
            f"<div class='status-label'>{item['label']}</div>"
            f"<div class='status-value {item.get('status_class', '')}'>"
            f"{item['value']}</div>"
            f"<div class='stat-sub'>{item.get('sub', '')}</div>"
            "</div>"
            for item in items
        ]
    )
    st.markdown(f"<div class='dashboard-grid'>{cards}</div>", unsafe_allow_html=True)

def get_db_conn() -> sqlite3.Connection:
    db_path = os.getenv("APP_DB_PATH", "app.db")
    conn = sqlite3.connect(db_path, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    return conn

def init_db() -> None:
    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            is_subscribed INTEGER DEFAULT 0,
            created_at TEXT NOT NULL
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            summary_json TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
        """
    )
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            session_id TEXT NOT NULL,
            event_name TEXT NOT NULL,
            details_json TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
        """
    )
    conn.commit()
    conn.close()

def hash_password(password: str) -> str:
    salt = os.getenv("APP_SALT", "trade-analyzer")
    return hashlib.sha256(f"{salt}:{password}".encode("utf-8")).hexdigest()

def get_user_by_email(email: str) -> Optional[Dict[str, Any]]:
    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute("SELECT * FROM users WHERE email = ?", (email.lower(),))
    row = cur.fetchone()
    conn.close()
    return dict(row) if row else None

def create_user(email: str, password: str) -> Dict[str, Any]:
    conn = get_db_conn()
    cur = conn.cursor()
    password_hash = hash_password(password)
    created_at = datetime.utcnow().isoformat()
    cur.execute(
        "INSERT INTO users (email, password_hash, is_subscribed, created_at) VALUES (?, ?, ?, ?)",
        (email.lower(), password_hash, 0, created_at),
    )
    conn.commit()
    user_id = cur.lastrowid
    conn.close()
    return {
        "id": user_id,
        "email": email.lower(),
        "password_hash": password_hash,
        "is_subscribed": 0,
        "created_at": created_at,
    }

def verify_user(email: str, password: str) -> Optional[Dict[str, Any]]:
    user = get_user_by_email(email)
    if not user:
        return None
    if user["password_hash"] != hash_password(password):
        return None
    return user

def set_subscription(user_id: int, is_subscribed: bool) -> None:
    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute(
        "UPDATE users SET is_subscribed = ? WHERE id = ?",
        (1 if is_subscribed else 0, user_id),
    )
    conn.commit()
    conn.close()

def save_run_summary(user_id: int, summary: Dict[str, Any]) -> None:
    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO runs (user_id, created_at, summary_json) VALUES (?, ?, ?)",
        (user_id, datetime.utcnow().isoformat(), json.dumps(summary)),
    )
    conn.commit()
    conn.close()

def load_run_summaries(user_id: int) -> pd.DataFrame:
    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute(
        "SELECT created_at, summary_json FROM runs WHERE user_id = ? ORDER BY created_at DESC",
        (user_id,),
    )
    rows = cur.fetchall()
    conn.close()
    data = []
    for row in rows:
        payload = json.loads(row["summary_json"])
        payload["created_at"] = row["created_at"]
        data.append(payload)
    return pd.DataFrame(data)

def log_event(event_name: str, user_id: Optional[int] = None, details: Optional[Dict[str, Any]] = None) -> None:
    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute(
        "INSERT INTO events (user_id, session_id, event_name, details_json, created_at) VALUES (?, ?, ?, ?, ?)",
        (
            user_id,
            st.session_state.get("session_id"),
            event_name,
            json.dumps(details or {}),
            datetime.utcnow().isoformat(),
        ),
    )
    conn.commit()
    conn.close()

def navigate_to_diagnostics(mistake_type: str) -> None:
    st.session_state["nav_tab"] = "Diagnostics"
    st.session_state["diagnostics_mistake_type"] = mistake_type

APP_PASS = os.getenv("APP_PASS", "")

if APP_PASS:
    code = st.sidebar.text_input("Access code", type="password")
    if code != APP_PASS:
        st.title("🔒 Trade Analyzer")
        if code:
            st.sidebar.error("Wrong access code.")
        else:
            st.sidebar.warning("Enter access code to continue.")
        st.stop()

# ---------------------------------------------------------------------------------------
init_db()

if "session_id" not in st.session_state:
    st.session_state["session_id"] = str(uuid.uuid4())
if "user" not in st.session_state:
    st.session_state["user"] = None
if "nav_tab" not in st.session_state:
    st.session_state["nav_tab"] = "Overview"



DATE_LINE_RE = re.compile(r"^\s*\d{1,2}/\d{1,2}/\d{4}\s*,")
HEADER_RE = re.compile(r"^\s*Run Date\s*,", re.IGNORECASE)
POST_MORTEM_TAGS = [
    "Entered too early / chased",
    "Held too long",
    "Wrong thesis / news",
    "Position size too big",
    "Didn’t cut at stop",
    "IV crush / earnings",
    "Rolled to avoid loss",
]
UNASSIGNED_TAG = "(unassigned)"
LEAK_RULES = {
    "Overtrading day": "Avoid high-frequency churn on red days.",
    "Big loss outlier": "Cap maximum loss per trade.",
    "Held losers longer": "Cut losers faster than winners.",
    "Gave back gains in ticker": "Stop trading tickers that flip to net negative.",
    "Averaged down (added to loser)": "No adding to losing positions.",
}

def load_csv_bytes(file_bytes: bytes) -> pd.DataFrame:
    """
    Tries to load CSV (comma or tab). If it looks like a Fidelity export with disclaimer footer,
    it strips non-table lines and retries.
    """
    text = file_bytes.decode("utf-8-sig", errors="replace")

    # Heuristic sep guess on first chunk
    sample = text[:4096]
    sep = "\t" if sample.count("\t") > sample.count(",") else ","

    def _read(t: str, sep_: str) -> pd.DataFrame:
        return pd.read_csv(io.StringIO(t), sep=sep_, engine="python", skip_blank_lines=True)

    def _read_flexible(t: str, sep_: str) -> pd.DataFrame:
        rows = []
        reader = csv.reader(io.StringIO(t), delimiter=sep_)
        for row in reader:
            if any(cell.strip() for cell in row):
                rows.append(row)
        if not rows:
            return pd.DataFrame()
        header = rows[0]
        width = len(header)
        fixed_rows = []
        for row in rows[1:]:
            if len(row) < width:
                row = row + [""] * (width - len(row))
            elif len(row) > width:
                row = row[: width - 1] + [sep_.join(row[width - 1 :])]
            fixed_rows.append(row)
        return pd.DataFrame(fixed_rows, columns=header)

    try:
        df = _read(text, sep)
        # If it parsed into 1 column and looks wrong, fall through to clean read
        if df.shape[1] == 1 and (str(df.columns[0]).lower().startswith("unnamed") or "," in str(df.columns[0])):
            raise ValueError("Single-column parse; attempting cleaned parse")
        return df
    except Exception:
        try:
            df = _read_flexible(text, sep)
            if df.shape[1] > 1:
                return df
        except Exception:
            pass
        # Clean: keep header + date-starting lines only
        lines = [ln for ln in text.splitlines() if ln.strip()]
        # find header
        header_idx = None
        for i, ln in enumerate(lines):
            if HEADER_RE.match(ln):
                header_idx = i
                break
        if header_idx is None:
            # Last resort: return the original error by re-raising
            raise

        keep = [lines[header_idx]]
        for ln in lines[header_idx + 1:]:
            if DATE_LINE_RE.match(ln):
                keep.append(ln)
            else:
                # stop at first non-date line after table begins (Fidelity disclaimer/footer)
                break

        cleaned = "\n".join(keep)
        # Fidelity is comma-separated
        try:
            df = _read(cleaned, ",")
        except Exception:
            df = _read_flexible(cleaned, ",")
        return df

# -------------------------
# Option parsing
# -------------------------

HUMAN_OPT_RE = re.compile(r"(?P<under>[A-Z]{1,6})\s+(?P<exp>\d{1,2}/\d{1,2}/\d{2,4})\s+(?P<strike>\d+(?:\.\d+)?)\s+(?P<right>[CP])\b")
OCC_OPT_RE = re.compile(r"^(?P<under>[A-Z ]{1,6})(?P<y>\d{2})(?P<m>\d{2})(?P<d>\d{2})(?P<right>[CP])(?P<k>\d{8})$")
FID_OPT_RE = re.compile(r"^-?(?P<under>[A-Z]{1,6})(?P<y>\d{2})(?P<m>\d{2})(?P<d>\d{2})(?P<right>[CP])(?P<k>\d{1,6})(?:\.\d+)?$")
FID_DESC_RE = re.compile(r"\b(?P<mon>JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)\s+(?P<day>\d{1,2})\s+(?P<yy>\d{2})\s+\$(?P<strike>\d+(?:\.\d+)?)\b", re.IGNORECASE)
MON_MAP = {"JAN":1,"FEB":2,"MAR":3,"APR":4,"MAY":5,"JUN":6,"JUL":7,"AUG":8,"SEP":9,"OCT":10,"NOV":11,"DEC":12}
STOCK_RE = re.compile(r"^[A-Z]{1,6}(?:\.[A-Z]{1,2})?$")

def _parse_mmddyy(s: str) -> Optional[pd.Timestamp]:
    dt = pd.to_datetime(str(s).strip(), errors="coerce")
    if pd.isna(dt):
        return None
    return pd.Timestamp(dt.date())

def parse_occ(symbol_text: str) -> Optional[Tuple[str, pd.Timestamp, float, str]]:
    t = str(symbol_text).strip().replace(" ", "")
    m = OCC_OPT_RE.match(t)
    if not m:
        return None
    under = m.group("under").strip()
    year = 2000 + int(m.group("y"))
    exp = pd.Timestamp(datetime(year, int(m.group("m")), int(m.group("d"))).date())
    right = m.group("right")
    strike = int(m.group("k")) / 1000.0
    return under, exp, float(strike), right

def parse_fidelity_symbol(sym: str) -> Optional[Tuple[str, pd.Timestamp, float, str]]:
    t = str(sym).strip().upper()
    m = FID_OPT_RE.match(t)
    if not m:
        return None
    under = m.group("under")
    year = 2000 + int(m.group("y"))
    exp = pd.Timestamp(datetime(year, int(m.group("m")), int(m.group("d"))).date())
    right = m.group("right")
    strike = float(m.group("k"))
    return under, exp, strike, right

def parse_fidelity_description(desc: str) -> Optional[Tuple[pd.Timestamp, float]]:
    t = str(desc).upper()
    m = FID_DESC_RE.search(t)
    if not m:
        return None
    mon = MON_MAP.get(m.group("mon").upper())
    day = int(m.group("day"))
    year = 2000 + int(m.group("yy"))
    exp = pd.Timestamp(datetime(year, mon, day).date())
    strike = float(m.group("strike"))
    return exp, strike

def parse_option(underlying_hint: Optional[str], symbol_text: str, desc_text: Optional[str]) -> Optional[Tuple[str, pd.Timestamp, float, str]]:
    out = parse_occ(symbol_text)
    if out:
        return out
    out = parse_fidelity_symbol(symbol_text)
    if out:
        return out
    m = HUMAN_OPT_RE.search(str(symbol_text).upper())
    if m:
        exp = _parse_mmddyy(m.group("exp"))
        if exp is None:
            return None
        return m.group("under"), exp, float(m.group("strike")), m.group("right")
    if desc_text is not None:
        td = str(desc_text).upper()
        right = "C" if "CALL" in td else ("P" if "PUT" in td else None)
        if right and underlying_hint:
            exs = parse_fidelity_description(td)
            if exs:
                exp, strike = exs
                return str(underlying_hint).strip().upper(), exp, float(strike), right
    return None

def looks_like_stock(symbol_text: str) -> bool:
    return bool(STOCK_RE.match(str(symbol_text).strip().upper()))

# -------------------------
# FIFO matching
# -------------------------
@dataclass
class Lot:
    qty: int
    price: float
    fees: float
    date: pd.Timestamp
    side: str
    multiplier: float

def fifo_match(trades: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    buys: List[Lot] = []
    sells: List[Lot] = []
    closed = []

    for _, r in trades.iterrows():
        lot = Lot(
            qty=int(r["qty"]),
            price=float(r["price"]),
            fees=float(r["fees"]),
            date=r["trade_dt"],
            side=r["side"],
            multiplier=float(r["multiplier"]),
        )
        if lot.side == "BUY":
            buys.append(lot)
        else:
            sells.append(lot)

        def match_once():
            nonlocal buys, sells, closed
            if not buys or not sells:
                return False
            b, s = buys[0], sells[0]
            mqty = min(b.qty, s.qty)
            multiplier = float(b.multiplier if b.multiplier is not None else s.multiplier)

            gross = (s.price - b.price) * mqty * multiplier
            b_fee = b.fees * (mqty / b.qty) if b.qty else 0.0
            s_fee = s.fees * (mqty / s.qty) if s.qty else 0.0
            net = gross - (b_fee + s_fee)

            closed.append({
                "buy_date": b.date,
                "sell_date": s.date,
                "close_date": max(b.date, s.date),
                "qty": mqty,
                "buy_price": b.price,
                "sell_price": s.price,
                "gross_pnl": gross,
                "fees": (b_fee + s_fee),
                "net_pnl": net
            })

            b.qty -= mqty; s.qty -= mqty
            b.fees -= b_fee; s.fees -= s_fee
            if b.qty == 0: buys.pop(0)
            if s.qty == 0: sells.pop(0)
            return True

        while match_once():
            pass

    open_rows = []
    for lot in buys:
        open_rows.append({
            "side": "BUY",
            "qty": lot.qty,
            "price": lot.price,
            "fees": lot.fees,
            "date": lot.date,
            "multiplier": lot.multiplier,
        })
    for lot in sells:
        open_rows.append({
            "side": "SELL",
            "qty": lot.qty,
            "price": lot.price,
            "fees": lot.fees,
            "date": lot.date,
            "multiplier": lot.multiplier,
        })

    return pd.DataFrame(closed), pd.DataFrame(open_rows)

# -------------------------
# Mistake detector helpers
# -------------------------
def _contract_key(df: pd.DataFrame) -> pd.Series:
    expiry = pd.to_datetime(df["expiry"], errors="coerce").dt.date.astype(str)
    strike = pd.to_numeric(df["strike"], errors="coerce").astype(str)
    right = df["right"].astype(str).fillna("STOCK")
    base = df["underlying"].astype(str)
    stock_mask = right.str.upper().eq("STOCK")
    opt_key = base + "|" + expiry + "|" + strike + "|" + right
    return opt_key.where(~stock_mask, base + "|STOCK")

def detect_mistakes(closed_trades: pd.DataFrame, fills: Optional[pd.DataFrame]) -> Dict[str, pd.DataFrame]:
    if closed_trades.empty:
        empty = pd.DataFrame(columns=[
            "mistake_type", "severity", "close_date", "underlying", "contract_key",
            "trade_id", "details", "net_pnl"
        ])
        return {
            "mistakes_df": empty,
            "summary_by_type": empty,
            "summary_by_week": empty,
            "examples_by_type": empty,
            "held_tables": {},
        }

    base = closed_trades.copy()
    base["close_date"] = pd.to_datetime(base["close_date"], errors="coerce")
    base["open_date"] = pd.to_datetime(base["open_date"], errors="coerce")
    base["trade_date"] = base["close_date"].dt.date
    base["holding_days"] = (base["close_date"] - base["open_date"]).dt.total_seconds() / (24 * 3600)
    base["contract_key"] = _contract_key(base)
    base["net_pnl"] = pd.to_numeric(base["net_pnl"], errors="coerce").fillna(0.0)

    mistakes: List[Dict[str, object]] = []

    # Rule 1: Overtrading day
    day_stats = (base.groupby("trade_date", as_index=False)
                 .agg(day_trade_count=("net_pnl", "size"), day_pnl=("net_pnl", "sum")))
    if not day_stats.empty:
        threshold = float(np.percentile(day_stats["day_trade_count"], 80))
        over = day_stats[(day_stats["day_trade_count"] >= threshold) & (day_stats["day_pnl"] < 0)]
        for _, row in over.iterrows():
            mistakes.append({
                "mistake_type": "Overtrading day",
                "severity": abs(float(row["day_pnl"])),
                "close_date": pd.Timestamp(row["trade_date"]),
                "underlying": "MULTI",
                "contract_key": "MULTI",
                "trade_id": str(row["trade_date"]),
                "details": f"{int(row['day_trade_count'])} closes, day PnL {row['day_pnl']:.2f}",
                "net_pnl": float(row["day_pnl"]),
            })

    # Rule 2: Big loss outlier
    losers = base[base["net_pnl"] < 0]
    winners = base[base["net_pnl"] > 0]
    if not losers.empty:
        base_loss = float(np.median(np.abs(losers["net_pnl"])))
        avg_win = float(winners["net_pnl"].mean()) if not winners.empty else 0.0
        threshold = max(3 * base_loss, 2 * avg_win) if avg_win > 0 else 3 * base_loss
        flagged = losers[np.abs(losers["net_pnl"]) >= threshold]
        for _, row in flagged.iterrows():
            mistakes.append({
                "mistake_type": "Big loss outlier",
                "severity": abs(float(row["net_pnl"])),
                "close_date": row["close_date"],
                "underlying": row["underlying"],
                "contract_key": row["contract_key"],
                "trade_id": str(row.get("trade_id", row.name)),
                "details": f"Loss {row['net_pnl']:.2f} exceeds threshold",
                "net_pnl": float(row["net_pnl"]),
            })

    # Rule 3: Held losers longer than winners
    held_tables = {}
    hold_ok = base["open_date"].notna().all()
    if hold_ok:
        winners_hold = base[base["net_pnl"] > 0]["holding_days"].dropna()
        losers_hold = base[base["net_pnl"] < 0]["holding_days"].dropna()
        if len(winners_hold) >= 5 and len(losers_hold) >= 5:
            med_w = float(np.median(winners_hold))
            med_l = float(np.median(losers_hold))
            if med_l > med_w * 1.25:
                last_close = base["close_date"].max()
                mistakes.append({
                    "mistake_type": "Held losers longer",
                    "severity": med_l - med_w,
                    "close_date": last_close,
                    "underlying": "MULTI",
                    "contract_key": "MULTI",
                    "trade_id": "global",
                    "details": f"Median hold losers {med_l:.1f}d vs winners {med_w:.1f}d",
                    "net_pnl": np.nan,
                })

        held_tables = {
            "losers": base[base["net_pnl"] < 0].sort_values("holding_days", ascending=False).head(10),
            "winners": base[base["net_pnl"] > 0].sort_values("holding_days", ascending=False).head(10),
        }

    # Rule 4: Gave back gains in ticker
    ticker = (base.groupby("underlying", as_index=False)
              .agg(ticker_net=("net_pnl", "sum"),
                   total_trades=("net_pnl", "size"),
                   has_wins=("net_pnl", lambda s: bool((s > 0).any())),
                   has_losses=("net_pnl", lambda s: bool((s < 0).any())),
                   last_close=("close_date", "max")))
    flagged = ticker[(ticker["ticker_net"] < 0) & ticker["has_wins"] & ticker["has_losses"] & (ticker["total_trades"] >= 5)]
    for _, row in flagged.iterrows():
        mistakes.append({
            "mistake_type": "Gave back gains in ticker",
            "severity": abs(float(row["ticker_net"])),
            "close_date": row["last_close"],
            "underlying": row["underlying"],
            "contract_key": "MULTI",
            "trade_id": str(row["underlying"]),
            "details": f"Net {row['ticker_net']:.2f} across {int(row['total_trades'])} trades with both wins & losses",
            "net_pnl": float(row["ticker_net"]),
        })

    # Rule 5: Averaged down losers (fills-based)
    if fills is not None and not fills.empty:
        fills_work = fills.copy()
        fills_work["contract_key"] = _contract_key(fills_work)
        fills_work["price"] = pd.to_numeric(fills_work["price"], errors="coerce").fillna(0.0)
        for contract_key, g in fills_work.sort_values("trade_dt").groupby("contract_key", sort=False):
            net_pos = 0
            avg_price = 0.0
            add_events = 0
            for _, r in g.iterrows():
                side = str(r["side"]).upper()
                qty = int(r["qty"])
                price = float(r["price"])
                before = net_pos
                if side == "BUY":
                    net_pos += qty
                    if before >= 0 and net_pos > before:
                        if before > 0 and price < avg_price:
                            add_events += 1
                        avg_price = ((avg_price * before) + price * qty) / max(net_pos, 1)
                    elif net_pos > 0:
                        avg_price = ((avg_price * max(before, 0)) + price * qty) / max(net_pos, 1)
                elif side == "SELL":
                    net_pos -= qty
                    if before <= 0 and net_pos < before:
                        if before < 0 and price > avg_price:
                            add_events += 1
                        avg_price = ((avg_price * abs(before)) + price * qty) / max(abs(net_pos), 1)
                    elif net_pos < 0:
                        avg_price = ((avg_price * abs(before)) + price * qty) / max(abs(net_pos), 1)
                if net_pos == 0:
                    avg_price = 0.0
            contract_net = float(base.loc[base["contract_key"] == contract_key, "net_pnl"].sum())
            if add_events >= 1 and contract_net < 0:
                parts = contract_key.split("|")
                mistakes.append({
                    "mistake_type": "Averaged down (added to loser)",
                    "severity": abs(contract_net),
                    "close_date": base.loc[base["contract_key"] == contract_key, "close_date"].max(),
                    "underlying": parts[0] if parts else "UNKNOWN",
                    "contract_key": contract_key,
                    "trade_id": contract_key,
                    "details": f"{add_events} adds at worse price, net {contract_net:.2f}",
                    "net_pnl": contract_net,
                })

    mistakes_df = pd.DataFrame(mistakes)
    if mistakes_df.empty:
        summary_by_type = pd.DataFrame(columns=["mistake_type", "count", "total_damage", "avg_damage"])
        summary_by_week = pd.DataFrame(columns=["week_key", "mistake_type", "count", "total_damage"])
        examples_by_type = pd.DataFrame(columns=mistakes_df.columns)
    else:
        mistakes_df["severity"] = pd.to_numeric(mistakes_df["severity"], errors="coerce").fillna(0.0)
        mistakes_df["close_date"] = pd.to_datetime(mistakes_df["close_date"], errors="coerce")
        summary_by_type = (mistakes_df.groupby("mistake_type", as_index=False)
                           .agg(count=("mistake_type", "size"),
                                total_damage=("severity", "sum"),
                                avg_damage=("severity", "mean"))
                           .sort_values("total_damage", ascending=False))
        iso = mistakes_df["close_date"].dt.isocalendar()
        week_key = iso["year"].astype(str) + "-W" + iso["week"].astype(str).str.zfill(2)
        summary_by_week = (mistakes_df.assign(week_key=week_key)
                           .groupby(["week_key", "mistake_type"], as_index=False)
                           .agg(count=("mistake_type", "size"),
                                total_damage=("severity", "sum")))
        examples_by_type = mistakes_df.sort_values("severity", ascending=False)

    return {
        "mistakes_df": mistakes_df,
        "summary_by_type": summary_by_type,
        "summary_by_week": summary_by_week,
        "examples_by_type": examples_by_type,
        "held_tables": held_tables,
    }

# -------------------------
# Daily P&L calendar helpers
# -------------------------
def compute_daily_pnl(closed_trades: pd.DataFrame) -> pd.DataFrame:
    if closed_trades.empty:
        return pd.DataFrame(columns=["trade_day", "day_pnl", "trade_count", "win_count", "loss_count"])

    base = closed_trades.copy()
    base["close_date"] = pd.to_datetime(base["close_date"], errors="coerce")
    base["net_pnl"] = pd.to_numeric(base["net_pnl"], errors="coerce").fillna(0.0)
    base = base[base["close_date"].notna()]
    base["trade_day"] = base["close_date"].dt.date

    daily = (base.groupby("trade_day", as_index=False)
             .agg(day_pnl=("net_pnl", "sum"),
                  trade_count=("net_pnl", "size"),
                  win_count=("net_pnl", lambda s: int((s > 0).sum())),
                  loss_count=("net_pnl", lambda s: int((s < 0).sum()))))
    return daily

def compute_equity_curve(closed_trades: pd.DataFrame) -> Tuple[pd.DataFrame, float]:
    if closed_trades.empty:
        empty = pd.DataFrame(columns=["trade_day", "daily_pnl", "cum_pnl", "peak", "drawdown"])
        return empty, 0.0

    base = closed_trades.copy()
    base["close_date"] = pd.to_datetime(base["close_date"], errors="coerce")
    base["net_pnl"] = pd.to_numeric(base["net_pnl"], errors="coerce").fillna(0.0)
    base = base[base["close_date"].notna()].copy()
    base["trade_day"] = base["close_date"].dt.date

    daily = (base.groupby("trade_day", as_index=False)
             .agg(daily_pnl=("net_pnl", "sum"))
             .sort_values("trade_day"))
    daily["cum_pnl"] = daily["daily_pnl"].cumsum()
    daily["peak"] = daily["cum_pnl"].cummax()
    daily["drawdown"] = daily["peak"] - daily["cum_pnl"]
    max_drawdown = float(daily["drawdown"].max()) if not daily.empty else 0.0
    return daily, max_drawdown

def _ensure_trade_id(trades: pd.DataFrame) -> pd.DataFrame:
    base = trades.copy()
    if "trade_id" not in base.columns:
        base["trade_id"] = base.index.astype(str)
    base["trade_id"] = base["trade_id"].fillna(base.index.astype(str)).astype(str)
    return base

def _prepare_closed_trades(trades: pd.DataFrame) -> pd.DataFrame:
    base = _ensure_trade_id(trades)
    base["row_order"] = np.arange(len(base))
    if "open_date" not in base.columns and {"buy_date", "sell_date"}.issubset(base.columns):
        base["open_date"] = base[["buy_date", "sell_date"]].min(axis=1)
    base["open_date"] = pd.to_datetime(base["open_date"], errors="coerce")
    base["close_date"] = pd.to_datetime(base["close_date"], errors="coerce")
    base["net_pnl"] = pd.to_numeric(base["net_pnl"], errors="coerce").fillna(0.0)
    base["trade_day"] = base["close_date"].dt.date
    base["contract_key"] = _contract_key(base)
    base["hold_days"] = (base["close_date"] - base["open_date"]).dt.total_seconds() / (24 * 3600)
    return base

def _compute_rule_defaults(trades: pd.DataFrame) -> Dict[str, float]:
    if trades.empty:
        return {"max_loss": 0.0, "red_day_limit": 0.0}
    losers = trades[trades["net_pnl"] < 0]
    max_loss = float(2 * np.median(np.abs(losers["net_pnl"]))) if not losers.empty else 0.0
    daily = compute_daily_pnl(trades)
    red_days = daily[daily["day_pnl"] < 0]
    red_day_limit = float(np.median(np.abs(red_days["day_pnl"]))) if not red_days.empty else 0.0
    return {"max_loss": max_loss, "red_day_limit": red_day_limit}

def _collapse_fill_events(fills: pd.DataFrame) -> pd.DataFrame:
    fills_work = fills.copy()
    fills_work["fill_dt"] = pd.to_datetime(fills_work["trade_dt"], errors="coerce")
    fills_work["side"] = fills_work["side"].astype(str).str.upper()
    fills_work["qty"] = pd.to_numeric(fills_work["qty"], errors="coerce").fillna(0.0)
    fills_work["price"] = pd.to_numeric(fills_work["price"], errors="coerce").fillna(0.0)
    fills_work["contract_key"] = _contract_key(fills_work)
    fills_work["signed_qty"] = np.where(fills_work["side"].eq("BUY"), fills_work["qty"], -fills_work["qty"])
    fills_work["fill_date"] = fills_work["fill_dt"].dt.date
    fills_work["row_order"] = np.arange(len(fills_work))

    if "order_id" in fills_work.columns and fills_work["order_id"].notna().any():
        grouped = (fills_work.groupby(["contract_key", "order_id", "side"], as_index=False)
                   .agg(signed_qty=("signed_qty", "sum"),
                        fill_dt=("fill_dt", "min"),
                        row_order=("row_order", "min")))
    else:
        grouped = (fills_work.groupby(["contract_key", "fill_date", "side", "price"], as_index=False)
                   .agg(signed_qty=("signed_qty", "sum"),
                        fill_dt=("fill_dt", "min"),
                        row_order=("row_order", "min")))
    grouped = grouped.sort_values(["contract_key", "fill_dt", "row_order"])
    return grouped

def _merge_equity_curves(base_curve: pd.DataFrame, sim_curve: pd.DataFrame) -> pd.DataFrame:
    merged = pd.merge(
        base_curve[["trade_day", "daily_pnl"]],
        sim_curve[["trade_day", "daily_pnl"]],
        on="trade_day",
        how="outer",
        suffixes=("_baseline", "_simulated"),
    ).sort_values("trade_day")
    merged["daily_pnl_baseline"] = merged["daily_pnl_baseline"].fillna(0.0)
    merged["daily_pnl_simulated"] = merged["daily_pnl_simulated"].fillna(0.0)
    merged["baseline_cum_pnl"] = merged["daily_pnl_baseline"].cumsum()
    merged["simulated_cum_pnl"] = merged["daily_pnl_simulated"].cumsum()
    merged["baseline_peak"] = merged["baseline_cum_pnl"].cummax()
    merged["simulated_peak"] = merged["simulated_cum_pnl"].cummax()
    merged["baseline_drawdown"] = merged["baseline_peak"] - merged["baseline_cum_pnl"]
    merged["simulated_drawdown"] = merged["simulated_peak"] - merged["simulated_cum_pnl"]
    return merged

def simulate_rules(
    closed_trades: pd.DataFrame,
    fills: Optional[pd.DataFrame],
    params: Optional[Dict[str, float]] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, float], Dict[str, Any]]:
    if closed_trades.empty:
        empty = closed_trades.copy()
        empty["sim_net_pnl"] = []
        empty["rules_applied"] = []
        return empty, pd.DataFrame(), {"baseline_pnl": 0.0, "max_drawdown": 0.0}, {
            "simulated_pnl": 0.0,
            "max_drawdown": 0.0,
            "rule_summaries": {},
            "equity_curve": pd.DataFrame(),
        }

    base = _prepare_closed_trades(closed_trades)
    base["baseline_net_pnl"] = base["net_pnl"]
    base["sim_net_pnl"] = base["net_pnl"]
    base["rules_applied"] = [[] for _ in range(len(base))]

    overrides: List[Dict[str, object]] = []
    rule_summaries: Dict[str, Dict[str, object]] = {}
    params = params or {}

    # Rule 1: No Averaging Down
    rule1_name = "Rule 1 — No Averaging Down"
    if fills is None or fills.empty:
        rule_summaries["rule1"] = {
            "name": rule1_name,
            "status": "unavailable",
            "message": "Rule 1 unavailable: requires fills (BUY/SELL events).",
            "violations": 0,
            "baseline_cost": 0.0,
            "simulated_savings": 0.0,
            "affected_trades": pd.DataFrame(),
        }
    else:
        events = _collapse_fill_events(fills)
        add_counts: Dict[str, int] = {}
        for contract_key, g in events.groupby("contract_key", sort=False):
            signed_qty = g["signed_qty"].astype(float)
            pos_before = signed_qty.cumsum().shift(1).fillna(0.0)
            pos_after = pos_before + signed_qty
            same_sign = np.sign(pos_before) == np.sign(pos_after)
            add_events = (pos_before != 0) & same_sign & (pos_after.abs() > pos_before.abs())
            add_counts[contract_key] = int(add_events.sum())

        base_contract_net = base.groupby("contract_key")["net_pnl"].sum()
        violating_contracts = [
            ck for ck, count in add_counts.items()
            if count >= 1 and base_contract_net.get(ck, 0.0) < 0
        ]
        for contract_key in violating_contracts:
            add_count = add_counts.get(contract_key, 0)
            entry_units = 1 + add_count
            contract_rows = base[base["contract_key"] == contract_key].copy()
            baseline_contract_net = float(contract_rows["net_pnl"].sum())
            simulated_contract_net = baseline_contract_net / entry_units if entry_units else baseline_contract_net
            abs_sum = float(contract_rows["net_pnl"].abs().sum())
            if abs_sum == 0:
                weights = np.repeat(1 / len(contract_rows), len(contract_rows))
            else:
                weights = contract_rows["net_pnl"].abs() / abs_sum
            for idx, weight in zip(contract_rows.index, weights):
                prev = float(base.at[idx, "sim_net_pnl"])
                new_value = float(simulated_contract_net * weight)
                base.at[idx, "sim_net_pnl"] = new_value
                if prev != new_value:
                    base.at[idx, "rules_applied"].append(rule1_name)
                    overrides.append({
                        "trade_id": base.at[idx, "trade_id"],
                        "rule_name": rule1_name,
                        "baseline_net_pnl": prev,
                        "simulated_net_pnl": new_value,
                        "delta": new_value - prev,
                        "override_reason": "Scaled by entry units (approximation)",
                    })

        affected = base[base["contract_key"].isin(violating_contracts)].copy()
        if not affected.empty:
            affected["override_reason"] = "Scaled by entry units (approximation)"
        baseline_cost = float(affected["baseline_net_pnl"].sum()) if not affected.empty else 0.0
        simulated_cost = float(affected["sim_net_pnl"].sum()) if not affected.empty else 0.0
        rule_summaries["rule1"] = {
            "name": rule1_name,
            "status": "active",
            "message": "Approximation by scaling to remove adds.",
            "violations": len(violating_contracts),
            "baseline_cost": baseline_cost,
            "simulated_savings": baseline_cost - simulated_cost,
            "affected_trades": affected,
        }

    # Rule 2: Max Loss Per Trade Cap
    rule2_name = "Rule 2 — Max Loss Per Trade Cap"
    max_loss = params.get("max_loss")
    losers = base[base["net_pnl"] < 0]
    if max_loss is None:
        max_loss = float(2 * np.median(np.abs(losers["net_pnl"]))) if not losers.empty else 0.0
    if max_loss <= 0:
        rule_summaries["rule2"] = {
            "name": rule2_name,
            "status": "unavailable",
            "message": "Rule 2 unavailable: no losing trades to set a cap.",
            "violations": 0,
            "baseline_cost": 0.0,
            "simulated_savings": 0.0,
            "affected_trades": pd.DataFrame(),
        }
    else:
        violating = base[base["net_pnl"] < -max_loss].copy()
        for idx in violating.index:
            prev = float(base.at[idx, "sim_net_pnl"])
            new_value = float(-max_loss)
            base.at[idx, "sim_net_pnl"] = new_value
            if prev != new_value:
                base.at[idx, "rules_applied"].append(rule2_name)
                overrides.append({
                    "trade_id": base.at[idx, "trade_id"],
                    "rule_name": rule2_name,
                    "baseline_net_pnl": prev,
                    "simulated_net_pnl": new_value,
                    "delta": new_value - prev,
                    "override_reason": f"Clamped to -${max_loss:,.2f}",
                })

        baseline_cost = float(violating["baseline_net_pnl"].sum()) if not violating.empty else 0.0
        simulated_cost = float(base.loc[violating.index, "sim_net_pnl"].sum()) if not violating.empty else 0.0
        affected = base.loc[violating.index].copy()
        if not affected.empty:
            affected["override_reason"] = f"Clamped to -${max_loss:,.2f}"
        rule_summaries["rule2"] = {
            "name": rule2_name,
            "status": "active",
            "message": f"Cap applied at -${max_loss:,.2f}.",
            "violations": len(violating),
            "baseline_cost": baseline_cost,
            "simulated_savings": baseline_cost - simulated_cost,
            "affected_trades": affected,
            "max_loss": max_loss,
        }

    # Rule 3: Stop Trading After Red Day Limit
    rule3_name = "Rule 3 — Stop Trading After Red Day Limit"
    red_day_limit = params.get("red_day_limit")
    daily = compute_daily_pnl(base)
    red_days = daily[daily["day_pnl"] < 0]
    if red_day_limit is None:
        red_day_limit = float(np.median(np.abs(red_days["day_pnl"]))) if not red_days.empty else 0.0
    if red_day_limit <= 0 or red_days.empty:
        rule_summaries["rule3"] = {
            "name": rule3_name,
            "status": "unavailable",
            "message": "Rule 3 unavailable: no red days to set a limit.",
            "violations": 0,
            "baseline_cost": 0.0,
            "simulated_savings": 0.0,
            "affected_trades": pd.DataFrame(),
        }
    else:
        trigger_days = daily[daily["day_pnl"] <= -red_day_limit]["trade_day"].tolist()
        skipped_indices: List[int] = []
        for trade_day in trigger_days:
            day_trades = base[base["trade_day"] == trade_day].sort_values(["close_date", "row_order"])
            day_trades["cum_pnl"] = day_trades["net_pnl"].cumsum()
            hit = day_trades[day_trades["cum_pnl"] <= -red_day_limit]
            if hit.empty:
                continue
            cutoff_idx = hit.index[0]
            after = day_trades.loc[day_trades.index > cutoff_idx]
            skipped_indices.extend(after.index.tolist())

        skipped = base.loc[sorted(set(skipped_indices))].copy()
        for idx in skipped.index:
            prev = float(base.at[idx, "sim_net_pnl"])
            new_value = 0.0
            base.at[idx, "sim_net_pnl"] = new_value
            if prev != new_value:
                base.at[idx, "rules_applied"].append(rule3_name)
                overrides.append({
                    "trade_id": base.at[idx, "trade_id"],
                    "rule_name": rule3_name,
                    "baseline_net_pnl": prev,
                    "simulated_net_pnl": new_value,
                    "delta": new_value - prev,
                    "override_reason": "Skipped after red-day limit hit",
                })

        baseline_cost = float(skipped["baseline_net_pnl"].sum()) if not skipped.empty else 0.0
        simulated_cost = float(base.loc[skipped.index, "sim_net_pnl"].sum()) if not skipped.empty else 0.0
        affected = base.loc[skipped.index].copy()
        if not affected.empty:
            affected["override_reason"] = "Skipped after red-day limit hit"
        rule_summaries["rule3"] = {
            "name": rule3_name,
            "status": "active",
            "message": f"Limit applied at -${red_day_limit:,.2f} per day.",
            "violations": len(skipped),
            "baseline_cost": baseline_cost,
            "simulated_savings": baseline_cost - simulated_cost,
            "affected_trades": affected,
            "red_day_limit": red_day_limit,
            "trigger_days": trigger_days,
        }

    overrides_df = pd.DataFrame(overrides)
    if overrides_df.empty:
        overrides_df = pd.DataFrame(columns=[
            "trade_id", "rule_name", "baseline_net_pnl", "simulated_net_pnl", "delta", "override_reason"
        ])

    baseline_curve, baseline_max_dd = compute_equity_curve(base)
    sim_curve, sim_max_dd = compute_equity_curve(base.assign(net_pnl=base["sim_net_pnl"]))
    sim_equity_curve = _merge_equity_curves(baseline_curve, sim_curve)

    baseline_metrics = {
        "baseline_pnl": float(base["net_pnl"].sum()),
        "max_drawdown": float(baseline_max_dd),
    }
    sim_metrics = {
        "simulated_pnl": float(base["sim_net_pnl"].sum()),
        "max_drawdown": float(sim_max_dd),
        "rule_summaries": rule_summaries,
        "equity_curve": sim_equity_curve,
    }
    return base, overrides_df, baseline_metrics, sim_metrics

def render_rules_coach_overview(
    closed_trades: pd.DataFrame,
    fills: Optional[pd.DataFrame],
) -> None:
    st.markdown("#### If you followed 3 rules, here’s the improvement")
    if closed_trades.empty:
        st.info("Upload closed trades to run the Rules Coach simulator.")
        return

    prepared = _prepare_closed_trades(closed_trades)
    defaults = _compute_rule_defaults(prepared)

    c1, c2 = st.columns(2)
    with c1:
        max_loss = st.number_input(
            "MAX_LOSS ($)",
            min_value=0.0,
            value=float(st.session_state.get("rules_max_loss", defaults["max_loss"])),
            step=10.0,
            help="Rule 2 cap for worst per-trade loss.",
            key="rules_max_loss_overview",
        )
        st.session_state["rules_max_loss"] = max_loss
    with c2:
        red_day_limit = st.number_input(
            "RED_DAY_LIMIT ($)",
            min_value=0.0,
            value=float(st.session_state.get("rules_red_day_limit", defaults["red_day_limit"])),
            step=10.0,
            help="Rule 3 stop-trading limit for red days.",
            key="rules_red_day_limit_overview",
        )
        st.session_state["rules_red_day_limit"] = red_day_limit

    sim_trades, overrides_df, baseline_metrics, sim_metrics = simulate_rules(
        prepared,
        fills,
        params={"max_loss": max_loss, "red_day_limit": red_day_limit},
    )

    baseline_pnl = baseline_metrics["baseline_pnl"]
    sim_pnl = sim_metrics["simulated_pnl"]
    pnl_delta = sim_pnl - baseline_pnl
    base_dd = baseline_metrics["max_drawdown"]
    sim_dd = sim_metrics["max_drawdown"]
    dd_delta = base_dd - sim_dd

    render_stat_cards(
        [
            {"label": "Baseline realized P&L", "value": f"${baseline_pnl:,.2f}", "sub": "Closed trades only"},
            {"label": "Simulated realized P&L", "value": f"${sim_pnl:,.2f}", "sub": "After 3 rules"},
            {"label": "Delta", "value": f"${pnl_delta:,.2f}", "sub": "Improvement"},
            {"label": "Baseline max drawdown", "value": f"${base_dd:,.2f}", "sub": "Realized equity"},
            {"label": "Simulated max drawdown", "value": f"${sim_dd:,.2f}", "sub": "Rules applied"},
            {"label": "Drawdown reduction", "value": f"${dd_delta:,.2f}", "sub": "Improvement"},
        ]
    )

    st.caption("Simulations are deterministic approximations based on your realized trades; no market data is used.")
    st.caption("Rule 1 uses a scaling approximation when removing adds because per-unit P&L is unavailable.")

    rule_summaries = sim_metrics.get("rule_summaries", {})
    st.markdown("#### Rule cards")
    rule_cols = st.columns(3)
    for col, key in zip(rule_cols, ["rule1", "rule2", "rule3"]):
        summary = rule_summaries.get(key, {})
        with col:
            st.markdown(f"**{summary.get('name', key)}**")
            if summary.get("status") == "unavailable":
                st.warning(summary.get("message", "Rule unavailable."))
            else:
                st.markdown(f"Violations: **{summary.get('violations', 0)}**")
                st.markdown(f"Cost: **${summary.get('baseline_cost', 0.0):,.2f}**")
                st.markdown(f"Simulated savings: **${summary.get('simulated_savings', 0.0):,.2f}**")
                if summary.get("message"):
                    st.caption(summary["message"])

            if st.button("View affected trades", key=f"view-affected-{key}"):
                st.session_state["rules_coach_selected_rule"] = key

    selected_rule = st.session_state.get("rules_coach_selected_rule")
    if selected_rule and rule_summaries.get(selected_rule):
        summary = rule_summaries[selected_rule]
        affected = summary.get("affected_trades", pd.DataFrame()).copy()
        if affected.empty:
            st.info("No affected trades for this rule.")
        else:
            affected = affected.assign(
                simulated_net_pnl=affected["sim_net_pnl"],
                delta=affected["sim_net_pnl"] - affected["baseline_net_pnl"],
            )
            st.markdown("#### Affected trades")
            st.dataframe(
                affected[[
                    "trade_id",
                    "open_date",
                    "close_date",
                    "contract_key",
                    "baseline_net_pnl",
                    "simulated_net_pnl",
                    "delta",
                    "override_reason",
                ]].sort_values("close_date", ascending=False),
                use_container_width=True,
            )

def render_rules_coach_diagnostics(
    closed_trades: pd.DataFrame,
    fills: Optional[pd.DataFrame],
) -> None:
    st.markdown("### Rules Coach Details")
    if closed_trades.empty:
        st.info("Upload closed trades to run the Rules Coach simulator.")
        return

    prepared = _prepare_closed_trades(closed_trades)
    defaults = _compute_rule_defaults(prepared)

    c1, c2 = st.columns(2)
    with c1:
        max_loss = st.number_input(
            "MAX_LOSS ($)",
            min_value=0.0,
            value=float(st.session_state.get("rules_max_loss", defaults["max_loss"])),
            step=10.0,
            key="rules_max_loss_diagnostics",
        )
    with c2:
        red_day_limit = st.number_input(
            "RED_DAY_LIMIT ($)",
            min_value=0.0,
            value=float(st.session_state.get("rules_red_day_limit", defaults["red_day_limit"])),
            step=10.0,
            key="rules_red_day_limit_diagnostics",
        )
    st.session_state["rules_max_loss"] = max_loss
    st.session_state["rules_red_day_limit"] = red_day_limit

    sim_trades, overrides_df, baseline_metrics, sim_metrics = simulate_rules(
        prepared,
        fills,
        params={"max_loss": max_loss, "red_day_limit": red_day_limit},
    )
    sim_trades["rules_applied"] = sim_trades["rules_applied"].apply(lambda rules: ", ".join(rules) if rules else "—")

    st.caption("Simulations are deterministic approximations based on your realized trades; no market data is used.")
    st.markdown("#### Affected trades")
    affected = sim_trades[sim_trades["sim_net_pnl"] != sim_trades["baseline_net_pnl"]].copy()
    if affected.empty:
        st.info("No trades were modified by the rule simulation.")
    else:
        affected = affected.assign(delta=affected["sim_net_pnl"] - affected["baseline_net_pnl"])
        st.dataframe(
            affected[[
                "trade_id",
                "trade_day",
                "contract_key",
                "baseline_net_pnl",
                "sim_net_pnl",
                "delta",
                "rules_applied",
            ]].sort_values("trade_day", ascending=False),
            use_container_width=True,
        )

    st.markdown("#### Equity curve comparison")
    equity_curve = sim_metrics.get("equity_curve", pd.DataFrame())
    if not equity_curve.empty:
        st.line_chart(
            equity_curve.set_index("trade_day")[["baseline_cum_pnl", "simulated_cum_pnl"]],
        )
        st.markdown("#### Drawdown comparison")
        st.line_chart(
            equity_curve.set_index("trade_day")[["baseline_drawdown", "simulated_drawdown"]],
        )

    st.markdown("#### Simulation audit trail")
    if overrides_df.empty:
        st.info("No overrides were applied.")
    else:
        st.dataframe(overrides_df.sort_values("trade_id"), use_container_width=True)

def apply_day_filter(closed_trades: pd.DataFrame, selected_day: Optional[date]) -> pd.DataFrame:
    if closed_trades.empty or selected_day is None:
        return closed_trades
    base = closed_trades.copy()
    base["close_date"] = pd.to_datetime(base["close_date"], errors="coerce")
    return base[base["close_date"].dt.date == selected_day].copy()

def render_pnl_calendar(daily_pnl_df: pd.DataFrame, month_yyyy_mm: str) -> Optional[date]:
    if daily_pnl_df.empty:
        st.info("No daily P&L data available for the calendar.")
        return st.session_state.get("selected_day")

    st.session_state.setdefault("selected_day", None)

    year, month = (int(part) for part in month_yyyy_mm.split("-"))
    month_df = daily_pnl_df.copy()
    month_df["trade_day"] = pd.to_datetime(month_df["trade_day"], errors="coerce").dt.date
    month_df = month_df[month_df["trade_day"].apply(lambda d: pd.notna(d) and d.month == month and d.year == year)]

    month_pnl = float(month_df["day_pnl"].sum()) if not month_df.empty else 0.0
    green_days = int((month_df["day_pnl"] > 0).sum()) if not month_df.empty else 0
    red_days = int((month_df["day_pnl"] < 0).sum()) if not month_df.empty else 0
    trading_days = int(month_df["trade_day"].nunique()) if not month_df.empty else 0
    avg_day_pnl = (month_pnl / trading_days) if trading_days else 0.0

    render_stat_cards(
        [
            {
                "label": "Month net P&L",
                "value": f"${month_pnl:,.0f}",
                "sub": f"Avg per trading day: ${avg_day_pnl:,.0f}",
            },
            {
                "label": "Green days",
                "value": f"{green_days:,}",
                "sub": f"Trading days: {trading_days:,}",
            },
            {
                "label": "Red days",
                "value": f"{red_days:,}",
                "sub": "Closed-day results",
            },
        ]
    )

    abs_max = float(month_df["day_pnl"].abs().max()) if not month_df.empty else 0.0
    first_weekday, num_days = calendar.monthrange(year, month)

    day_lookup = {
        row["trade_day"]: {
            "day_pnl": float(row["day_pnl"]),
            "trade_count": int(row["trade_count"]),
        }
        for _, row in month_df.iterrows()
    }

    st.markdown(
        """
        <style>
        .pnl-card-form {
            margin: 0;
        }
        .pnl-card {
            border-radius: 8px;
            padding: 0.45rem;
            min-height: 82px;
            border: 1px solid rgba(0, 0, 0, 0.06);
            box-shadow: 0 1px 1px rgba(0, 0, 0, 0.04);
            width: 100%;
            text-align: left;
            cursor: pointer;
            color: #111111;
        }
        .pnl-card .day {
            font-weight: 600;
            font-size: 0.85rem;
            margin-bottom: 0.25rem;
        }
        .pnl-card .pnl {
            font-size: 0.8rem;
            margin-bottom: 0.15rem;
        }
        .pnl-card .trades {
            font-size: 0.65rem;
            color: #222222;
        }
        .pnl-card:hover {
            border-color: rgba(0, 0, 0, 0.18);
            box-shadow: 0 2px 4px rgba(0, 0, 0, 0.1);
        }
        </style>
        """,
        unsafe_allow_html=True,
    )

    header_cols = st.columns(7)
    for col, label in zip(header_cols, ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]):
        with col:
            st.markdown(f"**{label}**")
    day_num = 1
    weeks = []
    week = [None] * first_weekday
    while day_num <= num_days:
        week.append(day_num)
        if len(week) == 7:
            weeks.append(week)
            week = []
        day_num += 1
    if week:
        week.extend([None] * (7 - len(week)))
        weeks.append(week)

    query_params = st.experimental_get_query_params()
    selected_day_param = query_params.get("selected_day", [None])[0]
    if selected_day_param:
        try:
            st.session_state["selected_day"] = date.fromisoformat(selected_day_param)
        except ValueError:
            pass

    for week in weeks:
        cols = st.columns(7)
        for idx, day in enumerate(week):
            with cols[idx]:
                if day is None:
                    st.markdown("&nbsp;", unsafe_allow_html=True)
                    continue

                day_date = date(year, month, day)
                data = day_lookup.get(day_date)
                day_pnl = data["day_pnl"] if data else 0.0
                trade_count = data["trade_count"] if data else 0

                if data is None:
                    bg_color = "rgba(230, 230, 230, 0.6)"
                else:
                    intensity = abs(day_pnl) / abs_max if abs_max else 0.0
                    alpha = 0.18 + (0.45 * min(1.0, intensity))
                    if day_pnl > 0:
                        bg_color = f"rgba(46, 204, 113, {alpha:.2f})"
                    elif day_pnl < 0:
                        bg_color = f"rgba(231, 76, 60, {alpha:.2f})"
                    else:
                        bg_color = "rgba(220, 220, 220, 0.6)"

                pnl_label = f"${day_pnl:,.0f}"
                day_param = day_date.isoformat()
                card_html = f"""
                <form class="pnl-card-form" action="" method="get">
                    <input type="hidden" name="selected_day" value="{day_param}" />
                    <button class="pnl-card" type="submit" style="background: {bg_color};">
                        <div class="day">{day}</div>
                        <div class="pnl">{pnl_label}</div>
                        <div class="trades">Trades: {trade_count}</div>
                    </button>
                </form>
                """
                st.markdown(card_html, unsafe_allow_html=True)

    return st.session_state.get("selected_day")

# -------------------------
# Broker detection + normalization
# -------------------------
def detect_broker(columns: List[str]) -> str:
    cols = {str(c).strip().lower() for c in columns}
    if {"run date", "action", "symbol", "quantity", "price"}.issubset(cols):
        return "Fidelity"
    if any(str(c).strip().lower() in {"date/time", "date", "timestamp", "trade date"} for c in columns) and ("symbol" in cols or "description" in cols):
        return "IBKR"
    return "Unknown"

def _colmap(df: pd.DataFrame) -> Dict[str, str]:
    return {str(c).strip().lower(): c for c in df.columns}

def _series(df: pd.DataFrame, name: str, default="") -> pd.Series:
    cmap = _colmap(df)
    key = str(name).strip().lower()
    col = cmap.get(key)
    if col is None:
        for k, v in cmap.items():
            if key in k or k in key:
                col = v
                break
    if col is None:
        return pd.Series([default] * len(df), index=df.index)
    return df[col]

def normalize_fidelity(df: pd.DataFrame, include_stocks: bool) -> pd.DataFrame:
    w = df.copy()

    run_date = _series(w, "Run Date", default=np.nan)
    w["trade_dt"] = pd.to_datetime(run_date, errors="coerce")
    w = w.dropna(subset=["trade_dt"])

    action = _series(w, "Action", default="").astype(str).str.upper()
    qty_raw = pd.to_numeric(_series(w, "Quantity", default=0.0), errors="coerce").fillna(0.0)

    w["side"] = np.where(action.str.contains("SOLD"), "SELL",
                 np.where(action.str.contains("BOUGHT"), "BUY",
                 np.where(qty_raw < 0, "SELL", "BUY")))
    w["qty"] = qty_raw.abs().astype(int)

    w["price"] = pd.to_numeric(_series(w, "Price", default=np.nan), errors="coerce")
    w = w.dropna(subset=["price"])

    comm = pd.to_numeric(_series(w, "Commission", default=0.0), errors="coerce").fillna(0.0)
    fees = pd.to_numeric(_series(w, "Fees", default=0.0), errors="coerce").fillna(0.0)
    w["fees"] = (comm + fees).astype(float)

    sym = _series(w, "Symbol", default="").astype(str)
    desc = _series(w, "Description", default="").astype(str)
    clean_sym = sym.str.replace("^-", "", regex=True).str.strip().str.upper()
    underlying_hint = clean_sym.str.extract(r"^([A-Z]{1,6})", expand=False)

    parsed = [parse_option(uh, s, d) for uh, s, d in zip(underlying_hint, sym, desc)]
    asset_type = []
    underlying = []
    expiry = []
    strike = []
    right = []
    for p, raw_sym in zip(parsed, clean_sym):
        if p is not None:
            u, e, k, r = p
            asset_type.append("Option")
            underlying.append(u)
            expiry.append(e)
            strike.append(k)
            right.append(r)
        elif include_stocks and looks_like_stock(raw_sym):
            asset_type.append("Stock")
            underlying.append(raw_sym)
            expiry.append(pd.NaT)
            strike.append(np.nan)
            right.append("Stock")
        else:
            asset_type.append(None)
            underlying.append(None)
            expiry.append(pd.NaT)
            strike.append(np.nan)
            right.append(None)

    ok = [a is not None for a in asset_type]
    w = w[ok].copy()
    w["asset_type"] = [a for a in asset_type if a is not None]
    w["underlying"] = [u for u, a in zip(underlying, asset_type) if a is not None]
    w["expiry"] = [e for e, a in zip(expiry, asset_type) if a is not None]
    w["strike"] = [k for k, a in zip(strike, asset_type) if a is not None]
    w["right"] = [r for r, a in zip(right, asset_type) if a is not None]

    w["multiplier"] = np.where(w["asset_type"] == "Stock", 1.0, 100.0)

    w["contract_id"] = np.where(
        w["asset_type"] == "Stock",
        w["underlying"].astype(str) + "|STOCK",
        w["underlying"].astype(str) + "|" + w["expiry"].astype(str) + "|" +
        w["strike"].astype(str) + "|" + w["right"].astype(str),
    )
    return w.sort_values("trade_dt").reset_index(drop=True)

def normalize_ibkr(df: pd.DataFrame, mapping: Dict[str, str], include_stocks: bool) -> pd.DataFrame:
    w = df.copy()

    w["trade_dt"] = pd.to_datetime(w[mapping["datetime"]], errors="coerce")
    w = w.dropna(subset=["trade_dt"])

    side_raw = w[mapping["side"]].astype(str).str.upper().str.strip()
    w["side"] = np.where(side_raw.str.contains("SELL"), "SELL", np.where(side_raw.str.contains("BUY"), "BUY", side_raw))

    w["qty"] = pd.to_numeric(w[mapping["qty"]], errors="coerce").fillna(0).astype(int).abs()
    w["price"] = pd.to_numeric(w[mapping["price"]], errors="coerce")
    w = w.dropna(subset=["price"])

    feescol = mapping.get("fees")
    w["fees"] = pd.to_numeric(w[feescol], errors="coerce").fillna(0.0).astype(float) if feescol else 0.0

    multcol = mapping.get("multiplier")
    w["multiplier"] = pd.to_numeric(w[multcol], errors="coerce").fillna(100.0).astype(float) if multcol else 100.0

    sym = w[mapping["symbol"]].astype(str)
    clean_sym = sym.str.strip().str.upper()
    parsed = [parse_option(None, s, None) for s in sym]
    asset_type = []
    underlying = []
    expiry = []
    strike = []
    right = []
    for p, raw_sym in zip(parsed, clean_sym):
        if p is not None:
            u, e, k, r = p
            asset_type.append("Option")
            underlying.append(u)
            expiry.append(e)
            strike.append(k)
            right.append(r)
        elif include_stocks and looks_like_stock(raw_sym):
            asset_type.append("Stock")
            underlying.append(raw_sym)
            expiry.append(pd.NaT)
            strike.append(np.nan)
            right.append("Stock")
        else:
            asset_type.append(None)
            underlying.append(None)
            expiry.append(pd.NaT)
            strike.append(np.nan)
            right.append(None)

    ok = [a is not None for a in asset_type]
    w = w[ok].copy()
    w["asset_type"] = [a for a in asset_type if a is not None]
    w["underlying"] = [u for u, a in zip(underlying, asset_type) if a is not None]
    w["expiry"] = [e for e, a in zip(expiry, asset_type) if a is not None]
    w["strike"] = [k for k, a in zip(strike, asset_type) if a is not None]
    w["right"] = [r for r, a in zip(right, asset_type) if a is not None]
    w["multiplier"] = np.where(w["asset_type"] == "Stock", 1.0, w["multiplier"])

    w["contract_id"] = np.where(
        w["asset_type"] == "Stock",
        w["underlying"].astype(str) + "|STOCK",
        w["underlying"].astype(str) + "|" + w["expiry"].astype(str) + "|" +
        w["strike"].astype(str) + "|" + w["right"].astype(str),
    )
    return w.sort_values("trade_dt").reset_index(drop=True)

def build_dedupe_key(df: pd.DataFrame) -> pd.Series:
    expiry = pd.to_datetime(df["expiry"], errors="coerce").dt.date.astype(str)
    strike = pd.to_numeric(df["strike"], errors="coerce").round(4).astype(str)
    price = pd.to_numeric(df["price"], errors="coerce").round(6).astype(str)
    fees = pd.to_numeric(df["fees"], errors="coerce").round(6).astype(str)
    mult = pd.to_numeric(df["multiplier"], errors="coerce").round(6).astype(str)
    trade_dt = pd.to_datetime(df["trade_dt"], errors="coerce").dt.strftime("%Y-%m-%d %H:%M:%S")
    parts = [
        trade_dt,
        df["side"].astype(str),
        df["qty"].astype(str),
        price,
        fees,
        mult,
        df["asset_type"].astype(str),
        df["underlying"].astype(str),
        expiry,
        strike,
        df["right"].astype(str),
    ]
    return pd.Series(["|".join(vals) for vals in zip(*parts)], index=df.index)

def compute_reconciliation_metrics(
    raw_rows: int,
    file_count: int,
    work: pd.DataFrame,
    open_positions: pd.DataFrame,
    duplicate_count: int,
    fee_columns: List[str],
) -> Dict[str, Any]:
    date_min = pd.to_datetime(work["trade_dt"], errors="coerce").min()
    date_max = pd.to_datetime(work["trade_dt"], errors="coerce").max()
    unmatched_legs = int((open_positions["asset_type"] == "Option").sum()) if not open_positions.empty else 0
    open_qty = int(open_positions["qty"].sum()) if not open_positions.empty else 0
    fees_included = len(fee_columns) > 0
    roll_confidence = "Basic OK" if unmatched_legs == 0 and open_qty == 0 else "Needs review"
    return {
        "raw_rows": raw_rows,
        "file_count": file_count,
        "date_range": (date_min, date_max),
        "duplicate_count": duplicate_count,
        "unmatched_legs": unmatched_legs,
        "open_qty": open_qty,
        "fees_included": fees_included,
        "fee_columns": fee_columns,
        "roll_confidence": roll_confidence,
    }

def build_share_image(
    daily_pnl: pd.DataFrame,
    month_str: str,
    summary: Dict[str, Any],
    top_leaks: List[Dict[str, Any]],
) -> bytes:
    year, month = [int(part) for part in month_str.split("-")]
    month_days = calendar.monthcalendar(year, month)
    pnl_map = {pd.to_datetime(r["trade_day"]).day: float(r["net_pnl"]) for _, r in daily_pnl.iterrows()}

    fig = plt.figure(figsize=(9, 6), dpi=150)
    fig.patch.set_facecolor("#0f172a")
    ax = fig.add_subplot(111)
    ax.axis("off")

    ax.text(0.02, 0.95, f"Trade Month Summary — {month_str}", fontsize=16, fontweight="bold", color="#f8fafc")
    ax.text(0.02, 0.90, f"Realized P&L: ${summary['realized_pnl']:,.2f}", fontsize=12, color="#38bdf8")
    ax.text(0.35, 0.90, f"Max drawdown: ${summary['max_drawdown']:,.2f}", fontsize=12, color="#f87171")

    for idx, leak in enumerate(top_leaks[:3]):
        ax.text(
            0.02,
            0.84 - idx * 0.04,
            f"Leak #{idx + 1}: {leak['mistake_type']} (${leak['total_damage']:,.2f})",
            fontsize=10,
            color="#fbbf24",
        )

    cal_top = 0.72
    cell_w = 0.12
    cell_h = 0.08
    weekdays = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    for i, wd in enumerate(weekdays):
        ax.text(0.02 + i * cell_w, cal_top + 0.05, wd, fontsize=9, color="#94a3b8")

    max_abs = max(abs(v) for v in pnl_map.values()) if pnl_map else 1
    for week_idx, week in enumerate(month_days):
        for day_idx, day in enumerate(week):
            x = 0.02 + day_idx * cell_w
            y = cal_top - week_idx * cell_h
            if day == 0:
                rect = plt.Rectangle((x, y), cell_w - 0.01, cell_h - 0.01, color="#1e293b", alpha=0.4)
                ax.add_patch(rect)
                continue
            pnl = pnl_map.get(day, 0.0)
            norm = pnl / max_abs if max_abs else 0
            if pnl > 0:
                color = (0.12, 0.8, 0.5, 0.35 + 0.4 * min(abs(norm), 1))
            elif pnl < 0:
                color = (0.96, 0.38, 0.38, 0.35 + 0.4 * min(abs(norm), 1))
            else:
                color = (0.3, 0.3, 0.4, 0.3)
            rect = plt.Rectangle((x, y), cell_w - 0.01, cell_h - 0.01, color=color)
            ax.add_patch(rect)
            ax.text(x + 0.01, y + 0.02, str(day), fontsize=8, color="#e2e8f0")
            if pnl != 0:
                ax.text(x + 0.01, y + 0.045, f"{pnl:,.0f}", fontsize=7, color="#e2e8f0")

    if summary.get("top_winner"):
        ax.text(0.65, 0.9, f"Top winner: {summary['top_winner']}", fontsize=10, color="#34d399")
    if summary.get("top_loser"):
        ax.text(0.65, 0.86, f"Top loser: {summary['top_loser']}", fontsize=10, color="#f87171")

    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight")
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()

# -------------------------
# UI
# -------------------------
st.title("Trade Analyzer Pro")
st.caption("Upload Fidelity or IBKR fills. The app reconciles, dedupes, and separates realized vs unrealized results.")

include_stocks = st.toggle(
    "Include stock trades in P&L calculations",
    value=False,
    help="When enabled, equity trades are included alongside options.",
)

uploaded_files = st.file_uploader("Upload CSVs (multi-file supported)", type=["csv", "txt"], accept_multiple_files=True)
if not uploaded_files:
    st.stop()

log_event("file_uploaded", details={"file_count": len(uploaded_files)}, user_id=st.session_state["user"]["id"] if st.session_state["user"] else None)

file_entries = []
raw_rows = 0
fee_columns: List[str] = []
file_errors = []
for idx, up in enumerate(uploaded_files):
    try:
        df = load_csv_bytes(up.getvalue())
    except Exception as e:
        file_errors.append(f"{up.name}: {e}")
        continue
    raw_rows += len(df)
    detected = detect_broker(list(df.columns))
    file_entries.append({"file": up, "df": df, "detected": detected, "broker": detected, "index": idx})

if file_errors:
    st.error("Some files could not be parsed:\n" + "\n".join(file_errors))
    log_event("error", details={"errors": file_errors}, user_id=st.session_state["user"]["id"] if st.session_state["user"] else None)
    st.stop()

def pick_like(cols: List[str], candidates: List[str]) -> str:
    for cand in candidates:
        for c in cols:
            if cand.lower() == str(c).lower():
                return c
    for cand in candidates:
        for c in cols:
            if cand.lower() in str(c).lower():
                return c
    return cols[0] if cols else ""

with st.expander("Upload details & column mapping", expanded=False):
    for entry in file_entries:
        st.markdown(f"**{entry['file'].name}**")
        cols = list(entry["df"].columns)
        choice = st.selectbox(
            "Broker format",
            ["Auto", "Fidelity", "IBKR"],
            index=0,
            key=f"broker-{entry['index']}",
        )
        broker = entry["detected"] if choice == "Auto" else choice
        entry["broker"] = broker
        st.caption(f"Detected: {entry['detected']} | Using: {broker}")
        if broker == "IBKR":
            c1, c2 = st.columns(2)
            with c1:
                col_datetime = st.selectbox(
                    "Trade datetime",
                    cols,
                    index=cols.index(pick_like(cols, ["Date/Time", "Date", "TradeDate", "Timestamp"])),
                    key=f"dt-{entry['index']}",
                )
                col_symbol = st.selectbox(
                    "Symbol/Description",
                    cols,
                    index=cols.index(pick_like(cols, ["Symbol", "Description", "Instrument"])),
                    key=f"sym-{entry['index']}",
                )
                col_side = st.selectbox(
                    "Side",
                    cols,
                    index=cols.index(pick_like(cols, ["Side", "Action", "Buy/Sell", "B/S"])),
                    key=f"side-{entry['index']}",
                )
                col_qty = st.selectbox(
                    "Quantity",
                    cols,
                    index=cols.index(pick_like(cols, ["Quantity", "Qty", "Size"])),
                    key=f"qty-{entry['index']}",
                )
            with c2:
                col_price = st.selectbox(
                    "Price",
                    cols,
                    index=cols.index(pick_like(cols, ["Price", "TradePrice", "FillPrice"])),
                    key=f"price-{entry['index']}",
                )
                col_fees = st.selectbox(
                    "Fees/Commission",
                    ["(none)"] + cols,
                    index=0,
                    key=f"fees-{entry['index']}",
                )
                col_mult = st.selectbox(
                    "Multiplier",
                    ["(none)"] + cols,
                    index=0,
                    key=f"mult-{entry['index']}",
                )
            entry["mapping"] = {
                "datetime": col_datetime,
                "symbol": col_symbol,
                "side": col_side,
                "qty": col_qty,
                "price": col_price,
                "fees": None if col_fees == "(none)" else col_fees,
                "multiplier": None if col_mult == "(none)" else col_mult,
            }
        elif broker == "Fidelity":
            entry["mapping"] = None
        else:
            st.warning("Unable to detect broker format. Select Fidelity or IBKR.")

normalized_frames = []
for entry in file_entries:
    broker = entry["broker"]
    df = entry["df"]
    if broker == "Fidelity":
        normalized = normalize_fidelity(df, include_stocks)
        fee_columns.extend([col for col in ["Commission", "Fees"] if col in df.columns])
    elif broker == "IBKR":
        mapping = entry.get("mapping")
        if mapping is None:
            continue
        normalized = normalize_ibkr(df, mapping, include_stocks)
        if mapping.get("fees"):
            fee_columns.append(mapping["fees"])
    else:
        continue
    normalized["source_file"] = entry["file"].name
    normalized_frames.append(normalized)

if not normalized_frames:
    st.warning("No trades parsed from these files.")
    st.stop()

work = pd.concat(normalized_frames, ignore_index=True)
work["dedupe_key"] = build_dedupe_key(work)
duplicate_count = int(work.duplicated("dedupe_key").sum())
dup_rows = work[work.duplicated("dedupe_key", keep="first")].copy()
work = work.drop_duplicates("dedupe_key", keep="first").drop(columns=["dedupe_key"]).reset_index(drop=True)

# FIFO per contract
closed_all = []
open_all = []
for cid, g in work.groupby("contract_id", sort=False):
    g2 = g[["trade_dt", "side", "qty", "price", "fees", "multiplier", "asset_type", "underlying", "expiry", "strike", "right"]].sort_values("trade_dt")
    closed_df, open_df = fifo_match(g2)
    if not closed_df.empty:
        closed_df["contract_id"] = cid
        closed_df["asset_type"] = g2["asset_type"].iloc[0]
        closed_df["underlying"] = g2["underlying"].iloc[0]
        closed_df["expiry"] = g2["expiry"].iloc[0]
        closed_df["strike"] = g2["strike"].iloc[0]
        closed_df["right"] = g2["right"].iloc[0]
        closed_all.append(closed_df)
    if not open_df.empty:
        open_df["contract_id"] = cid
        open_df["asset_type"] = g2["asset_type"].iloc[0]
        open_df["underlying"] = g2["underlying"].iloc[0]
        open_df["expiry"] = g2["expiry"].iloc[0]
        open_df["strike"] = g2["strike"].iloc[0]
        open_df["right"] = g2["right"].iloc[0]
        open_all.append(open_df)

closed = pd.concat(closed_all, ignore_index=True) if closed_all else pd.DataFrame()
openpos = pd.concat(open_all, ignore_index=True) if open_all else pd.DataFrame()

trades_sheet = closed.copy().reset_index(drop=True) if not closed.empty else pd.DataFrame()
if not trades_sheet.empty:
    trades_sheet["trade_id"] = trades_sheet.index.astype(str)
spy_trades = trades_sheet[trades_sheet["underlying"] == "SPY"].copy() if not trades_sheet.empty else pd.DataFrame()
spy_total = float(spy_trades["net_pnl"].sum()) if not spy_trades.empty else 0.0

log_event(
    "analysis_run",
    details={"closed_trades": len(trades_sheet), "open_positions": len(openpos)},
    user_id=st.session_state["user"]["id"] if st.session_state["user"] else None,
)

recon = compute_reconciliation_metrics(
    raw_rows=raw_rows,
    file_count=len(uploaded_files),
    work=work,
    open_positions=openpos,
    duplicate_count=duplicate_count,
    fee_columns=sorted(set(fee_columns)),
)

date_min, date_max = recon["date_range"]
if pd.isna(date_min) or pd.isna(date_max):
    date_range_label = "—"
else:
    date_range_label = f"{date_min.date()} → {date_max.date()}"

issues = []
if duplicate_count:
    for _, row in dup_rows.head(50).iterrows():
        issues.append({
            "issue_type": "Duplicate fill",
            "trade_dt": row["trade_dt"],
            "underlying": row["underlying"],
            "details": f"{row['side']} {row['qty']} @ {row['price']}",
        })
if not openpos.empty:
    for _, row in openpos.head(50).iterrows():
        issues.append({
            "issue_type": "Unmatched/open leg",
            "trade_dt": row["date"],
            "underlying": row["underlying"],
            "details": f"{row['side']} {row['qty']} @ {row['price']}",
        })

has_issues = duplicate_count > 0 or recon["unmatched_legs"] > 0 or recon["open_qty"] > 0
status_icon = "✅" if not has_issues else "⚠️"
status_class = "status-clean" if not has_issues else "status-warn"

st.markdown("### Reconciliation & Data Quality")
render_status_cards(
    [
        {"label": "Rows parsed", "value": f"{recon['raw_rows']:,}", "status_class": status_class},
        {"label": "Files uploaded", "value": f"{recon['file_count']:,}", "status_class": status_class},
        {
            "label": "Date range",
            "value": date_range_label,
            "status_class": status_class,
        },
        {"label": "Duplicates removed", "value": f"{recon['duplicate_count']:,}", "status_class": status_class},
        {"label": "Unmatched legs", "value": f"{recon['unmatched_legs']:,}", "status_class": status_class},
        {"label": "Open qty leftover", "value": f"{recon['open_qty']:,}", "status_class": status_class},
        {
            "label": "Fees included",
            "value": f"{'Yes' if recon['fees_included'] else 'No'} ({', '.join(recon['fee_columns']) or '—'})",
            "status_class": "status-clean" if recon["fees_included"] else "status-warn",
        },
        {
            "label": "Roll detection",
            "value": recon["roll_confidence"],
            "status_class": status_class,
        },
    ]
)

if has_issues:
    st.warning(f"{status_icon} Issues detected — review before trusting results.")

with st.expander("View issues", expanded=has_issues):
    if issues:
        st.dataframe(pd.DataFrame(issues).head(50), use_container_width=True)
    else:
        st.success("No reconciliation issues detected.")

nav_tabs = ["Overview", "Journal", "Options", "Diagnostics", "Export", "Settings"]
nav_index = nav_tabs.index(st.session_state.get("nav_tab", "Overview"))
nav_selection = st.radio(
    "Navigation",
    nav_tabs,
    index=nav_index,
    horizontal=True,
    key="nav_tab",
    label_visibility="collapsed",
)
if st.session_state.get("last_nav") != nav_selection:
    st.session_state["last_nav"] = nav_selection
    log_event("page_view", details={"tab": nav_selection}, user_id=st.session_state["user"]["id"] if st.session_state["user"] else None)

mistake_output = {"mistakes_df": pd.DataFrame(), "summary_by_type": pd.DataFrame(), "summary_by_week": pd.DataFrame(),
                  "examples_by_type": pd.DataFrame(), "held_tables": {}}
if not trades_sheet.empty:
    base_mistakes = trades_sheet.copy()
    base_mistakes["open_date"] = base_mistakes[["buy_date", "sell_date"]].min(axis=1)
    base_mistakes["close_date"] = base_mistakes[["buy_date", "sell_date"]].max(axis=1)
    mistake_output = detect_mistakes(base_mistakes, work)

daily_pnl = pd.DataFrame()
if not trades_sheet.empty:
    ts = trades_sheet.copy()
    ts["open_date"] = ts[["buy_date", "sell_date"]].min(axis=1)
    ts["close_date"] = ts[["buy_date", "sell_date"]].max(axis=1)
    ts["hold_days"] = (ts["close_date"] - ts["open_date"]).dt.total_seconds() / (24 * 3600)
    ts["position_type"] = np.where(ts["sell_date"] < ts["buy_date"], "SHORT", "LONG")
    daily_pnl = compute_daily_pnl(ts)

mark_data_available = False

if nav_selection == "Overview":
    st.markdown("### Overview")
    if trades_sheet.empty:
        st.info("No closed trades to display yet.")
    else:
        pnl_total = float(trades_sheet["net_pnl"].sum())
        winrate = float((trades_sheet["net_pnl"] > 0).mean())
        wins = trades_sheet[trades_sheet["net_pnl"] > 0]["net_pnl"]
        losses = trades_sheet[trades_sheet["net_pnl"] < 0]["net_pnl"]
        profit_factor = wins.sum() / abs(losses.sum()) if not losses.empty else np.inf

        curve = trades_sheet.sort_values("close_date")[["close_date", "net_pnl"]].copy()
        curve["cum_net"] = curve["net_pnl"].cumsum()
        curve["peak"] = curve["cum_net"].cummax()
        curve["drawdown"] = curve["cum_net"] - curve["peak"]
        max_dd = float(curve["drawdown"].min()) if not curve.empty else 0.0
        avg_win = float(wins.mean()) if not wins.empty else 0.0
        avg_loss = float(losses.mean()) if not losses.empty else 0.0

        render_stat_cards(
            [
                {"label": "Realized P&L", "value": f"${pnl_total:,.2f}", "sub": "Closed trades only"},
                {"label": "Win rate", "value": f"{winrate*100:,.1f}%", "sub": "Realized only"},
                {"label": "Profit factor", "value": "∞" if np.isinf(profit_factor) else f"{profit_factor:,.2f}", "sub": "Realized only"},
                {"label": "Max drawdown", "value": f"${abs(max_dd):,.2f}", "sub": "Realized equity curve"},
                {"label": "# Trades", "value": f"{len(trades_sheet):,}", "sub": "Closed matches"},
                {"label": "Avg win / Avg loss", "value": f"${avg_win:,.0f} / ${avg_loss:,.0f}", "sub": "Realized only"},
            ]
        )

        render_rules_coach_overview(trades_sheet, work)

        c1, c2 = st.columns([2, 1])
        with c1:
            st.markdown("#### Equity curve + drawdown (realized)")
            st.line_chart(curve.set_index("close_date")[["cum_net", "drawdown"]])
        with c2:
            st.markdown("#### Top 3 leaks costing you money")
            summary_by_type = mistake_output["summary_by_type"]
            if summary_by_type.empty:
                st.info("No rule-based leaks detected yet.")
            else:
                top_leaks = summary_by_type.head(3).to_dict("records")
                for leak in top_leaks:
                    rule_text = LEAK_RULES.get(leak["mistake_type"], "Follow your risk rules.")
                    st.markdown(
                        f"**{leak['mistake_type']}**  \n"
                        f"Damage: ${leak['total_damage']:,.2f} • Count: {int(leak['count'])}  \n"
                        f"_Rule_: {rule_text}",
                    )
                    st.button(
                        "Show examples",
                        key=f"leak-{leak['mistake_type']}",
                        on_click=navigate_to_diagnostics,
                        args=(leak["mistake_type"],),
                    )

        st.markdown("#### Daily P&L calendar (realized)")
        if daily_pnl.empty:
            st.info("No closed trades available to build the daily calendar.")
        else:
            daily_pnl["month"] = pd.to_datetime(daily_pnl["trade_day"]).dt.to_period("M").astype(str)
            month_options = sorted(daily_pnl["month"].unique().tolist())
            default_month = pd.to_datetime(trades_sheet["close_date"]).max().strftime("%Y-%m")
            default_index = month_options.index(default_month) if default_month in month_options else len(month_options) - 1
            month_sel = st.selectbox("Month", options=month_options, index=default_index, key="overview-month")
            selected_day = render_pnl_calendar(daily_pnl, month_sel)
            if selected_day:
                st.caption(f"Trades closed on {selected_day:%Y-%m-%d}")

        st.markdown("#### Top tickers")
        by_ticker = (trades_sheet.groupby("underlying", as_index=False)
                     .agg(net_pnl=("net_pnl", "sum"))
                     .sort_values("net_pnl", ascending=False))
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("**Top 5 winners**")
            st.dataframe(by_ticker.head(5), use_container_width=True)
        with c2:
            st.markdown("**Top 5 losers**")
            st.dataframe(by_ticker.tail(5).sort_values("net_pnl"), use_container_width=True)

        if st.session_state["user"]:
            summary_payload = {
                "realized_pnl": pnl_total,
                "win_rate": winrate,
                "profit_factor": float(profit_factor) if np.isfinite(profit_factor) else None,
                "max_drawdown": abs(max_dd),
                "trade_count": len(trades_sheet),
                "avg_win": avg_win,
                "avg_loss": avg_loss,
                "top_winner": by_ticker.head(1)["underlying"].iloc[0] if not by_ticker.empty else None,
                "top_loser": by_ticker.tail(1)["underlying"].iloc[0] if not by_ticker.empty else None,
            }
            if st.button("Save run"):
                save_run_summary(st.session_state["user"]["id"], summary_payload)
                st.success("Run saved to your account.")
        else:
            st.info("Sign in to save runs and view history.")

elif nav_selection == "Journal":
    st.markdown("### Journal (realized trades)")
    if trades_sheet.empty:
        st.info("No closed trades to analyze.")
    else:
        journal_data = trades_sheet.copy()
        if not st.session_state["user"] or (st.session_state["user"] and not st.session_state["user"]["is_subscribed"]):
            cutoff = journal_data["close_date"].max() - pd.Timedelta(days=30)
            journal_data = journal_data[journal_data["close_date"] >= cutoff]
            st.caption("Free tier: showing the last 30 days. Upgrade to unlock full history.")

        journal_data["position_type"] = np.where(journal_data["sell_date"] < journal_data["buy_date"], "SHORT", "LONG")
        with st.expander("Filters", expanded=True):
            c1, c2, c3 = st.columns(3)
            with c1:
                under_sel = st.multiselect("Underlying", options=sorted(journal_data["underlying"].unique().tolist()),
                                           default=sorted(journal_data["underlying"].unique().tolist()))
            with c2:
                right_sel = st.multiselect("Type (C/P/Stock)", options=sorted(journal_data["right"].unique().tolist()),
                                           default=sorted(journal_data["right"].unique().tolist()))
            with c3:
                pos_sel = st.multiselect("Position (SHORT/LONG)", options=sorted(journal_data["position_type"].unique().tolist()),
                                         default=sorted(journal_data["position_type"].unique().tolist()))
        filtered = journal_data[
            journal_data["underlying"].isin(under_sel)
            & journal_data["right"].isin(right_sel)
            & journal_data["position_type"].isin(pos_sel)
        ].copy()
        if filtered.empty:
            st.warning("No trades match your filters.")
        else:
            st.dataframe(
                filtered.sort_values("close_date", ascending=False).head(200),
                use_container_width=True,
            )
            st.caption("Showing top 200 rows. Download full data from Export.")

elif nav_selection == "Options":
    if not st.session_state["user"] or (st.session_state["user"] and not st.session_state["user"]["is_subscribed"]):
        st.warning("Options view is part of the paid plan. Upgrade to unlock.")
    else:
        st.markdown("### Options & Open Positions")
        if openpos.empty:
            st.info("No open positions detected.")
        else:
            if mark_data_available:
                st.success("Mark data available — unrealized P&L shown separately.")
            else:
                st.warning("Unrealized P&L not available yet (no mark data).")
            st.dataframe(openpos.sort_values(["underlying", "expiry", "strike", "right", "side", "date"]), use_container_width=True)

elif nav_selection == "Diagnostics":
    if not st.session_state["user"] or (st.session_state["user"] and not st.session_state["user"]["is_subscribed"]):
        st.warning("Diagnostics are part of the paid plan. Upgrade to unlock.")
    else:
        st.markdown("### Diagnostics & Mistake Detector")
        render_rules_coach_diagnostics(trades_sheet, work)
        mistakes_df = mistake_output["mistakes_df"]
        summary_by_type = mistake_output["summary_by_type"]
        summary_by_week = mistake_output["summary_by_week"]
        examples_by_type = mistake_output["examples_by_type"]
        held_tables = mistake_output["held_tables"]

        if trades_sheet.empty:
            st.info("No closed trades to analyze for mistakes.")
        elif mistakes_df.empty:
            st.info("No mistakes detected with current rules.")
        else:
            st.dataframe(summary_by_type, use_container_width=True)
            if not summary_by_week.empty:
                top_types = summary_by_type.head(3)["mistake_type"].tolist()
                weekly = summary_by_week[summary_by_week["mistake_type"].isin(top_types)]
                weekly_pivot = weekly.pivot(index="week_key", columns="mistake_type", values="total_damage").fillna(0)
                st.line_chart(weekly_pivot)

            type_options = summary_by_type["mistake_type"].tolist()
            default_type = st.session_state.get("diagnostics_mistake_type")
            default_index = type_options.index(default_type) if default_type in type_options else 0
            type_sel = st.selectbox("Mistake type", type_options, index=default_index)
            examples = examples_by_type[examples_by_type["mistake_type"] == type_sel].head(20).copy()
            examples["date"] = examples["close_date"].dt.date
            st.dataframe(examples[["date", "underlying", "contract_key", "net_pnl", "severity", "details"]], use_container_width=True)

            if not examples.empty:
                ex_idx = st.selectbox(
                    "Inspect example row",
                    examples.index.tolist(),
                    format_func=lambda i: f"{examples.loc[i, 'date']} | {examples.loc[i, 'details']}",
                )
                selected = examples.loc[ex_idx]
                trades_view = trades_sheet.copy()
                trades_view["close_date"] = trades_view[["buy_date", "sell_date"]].max(axis=1)
                trades_view["contract_key"] = _contract_key(trades_view)

                if type_sel == "Overtrading day":
                    trades_view = trades_view[trades_view["close_date"].dt.date == selected["date"]]
                elif type_sel == "Gave back gains in ticker":
                    trades_view = trades_view[trades_view["underlying"] == selected["underlying"]]
                elif type_sel == "Big loss outlier":
                    trades_view = trades_view[trades_view["trade_id"] == selected["trade_id"]]
                elif type_sel == "Averaged down (added to loser)":
                    trades_view = trades_view[trades_view["contract_key"] == selected["contract_key"]]
                st.markdown("#### Related closed trades")
                st.dataframe(trades_view.sort_values("close_date"), use_container_width=True)

            if "losers" in held_tables and "winners" in held_tables:
                st.markdown("### Holding time contrast")
                c1, c2 = st.columns(2)
                with c1:
                    st.dataframe(held_tables["losers"][["close_date", "underlying", "contract_key", "net_pnl", "holding_days"]], use_container_width=True)
                with c2:
                    st.dataframe(held_tables["winners"][["close_date", "underlying", "contract_key", "net_pnl", "holding_days"]], use_container_width=True)

elif nav_selection == "Export":
    if not st.session_state["user"] or (st.session_state["user"] and not st.session_state["user"]["is_subscribed"]):
        st.warning("Export is part of the paid plan. Upgrade to unlock.")
    else:
        st.markdown("### Export")
        with st.expander("Parsed trades (raw)", expanded=False):
            st.dataframe(
                work[["trade_dt", "asset_type", "underlying", "expiry", "strike", "right", "side", "qty", "price", "fees"]].head(80),
                use_container_width=True,
            )

        if not trades_sheet.empty:
            with st.expander("Dashboard (realized)", expanded=False):
                dash = (trades_sheet.groupby("underlying", as_index=False)
                        .agg(net_pnl=("net_pnl", "sum"), gross_pnl=("gross_pnl", "sum"), fees=("fees", "sum"))
                        ).sort_values("net_pnl", ascending=False)
                st.dataframe(dash, use_container_width=True)

        with st.expander("Export Excel", expanded=True):
            fname = st.text_input("Output filename", "trade_analysis.xlsx")

            def to_excel_bytes():
                import io as _io
                from openpyxl.styles import PatternFill, Font
                from openpyxl.utils import get_column_letter

                buf = _io.BytesIO()

                dash = pd.DataFrame()
                if not trades_sheet.empty:
                    dash = (trades_sheet.groupby("underlying", as_index=False)
                            .agg(net_pnl=("net_pnl", "sum"), gross_pnl=("gross_pnl", "sum"), fees=("fees", "sum"))
                            ).sort_values("net_pnl", ascending=False)

                with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                    if not dash.empty:
                        dash.to_excel(writer, sheet_name="Dashboard", index=False)
                    if not trades_sheet.empty:
                        trades_sheet.to_excel(writer, sheet_name="Trades", index=False)
                    if not spy_trades.empty:
                        spy_trades.to_excel(writer, sheet_name="SPY", index=False)
                    if not openpos.empty:
                        openpos.to_excel(writer, sheet_name="OpenPositions", index=False)

                    work.to_excel(writer, sheet_name="RawParsed", index=False)

                    wb = writer.book

                    if "Trades" in wb.sheetnames:
                        ws = wb["Trades"]
                        headers = [cell.value for cell in ws[1]]
                        if "underlying" in headers:
                            idx = headers.index("underlying") + 1
                            fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                            for r in range(2, ws.max_row + 1):
                                if ws.cell(row=r, column=idx).value == "SPY":
                                    for c in range(1, ws.max_column + 1):
                                        ws.cell(row=r, column=c).fill = fill
                        for col in range(1, min(ws.max_column, 20) + 1):
                            ws.column_dimensions[get_column_letter(col)].width = 16

                    if "SPY" in wb.sheetnames:
                        ws = wb["SPY"]
                        last = ws.max_row + 2
                        ws.cell(row=last, column=1).value = "SPY net total (closed)"
                        ws.cell(row=last, column=2).value = spy_total
                        ws.cell(row=last, column=1).font = Font(bold=True)
                        ws.cell(row=last, column=2).font = Font(bold=True)

                buf.seek(0)
                return buf.getvalue()

            if st.button("Generate Excel"):
                st.download_button(
                    "Download Excel",
                    data=to_excel_bytes(),
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                log_event("export_clicked", details={"type": "excel"}, user_id=st.session_state["user"]["id"] if st.session_state["user"] else None)
                st.success("Generated Excel.")

        st.markdown("### Share my month (PNG)")
        if daily_pnl.empty:
            st.info("Need closed trades to generate a shareable month summary.")
        else:
            daily_pnl["month"] = pd.to_datetime(daily_pnl["trade_day"]).dt.to_period("M").astype(str)
            month_options = sorted(daily_pnl["month"].unique().tolist())
            month_sel = st.selectbox("Month to share", month_options, index=len(month_options) - 1, key="share-month")
            summary_by_type = mistake_output["summary_by_type"]
            top_leaks = summary_by_type.head(3).to_dict("records") if not summary_by_type.empty else []
            by_ticker = (trades_sheet.groupby("underlying", as_index=False)
                         .agg(net_pnl=("net_pnl", "sum"))
                         .sort_values("net_pnl", ascending=False))
            summary_payload = {
                "realized_pnl": float(trades_sheet["net_pnl"].sum()),
                "max_drawdown": abs(float(trades_sheet.sort_values("close_date")["net_pnl"].cumsum().sub(
                    trades_sheet.sort_values("close_date")["net_pnl"].cumsum().cummax()).min())),
                "top_winner": by_ticker.head(1)["underlying"].iloc[0] if not by_ticker.empty else None,
                "top_loser": by_ticker.tail(1)["underlying"].iloc[0] if not by_ticker.empty else None,
            }
            png_bytes = build_share_image(daily_pnl[daily_pnl["month"] == month_sel], month_sel, summary_payload, top_leaks)
            st.download_button(
                "Download shareable PNG",
                data=png_bytes,
                file_name=f"trade-summary-{month_sel}.png",
                mime="image/png",
            )

elif nav_selection == "Settings":
    st.markdown("### Settings")
    if st.session_state["user"]:
        st.success(f"Signed in as {st.session_state['user']['email']}")
        if st.button("Sign out"):
            st.session_state["user"] = None
            st.experimental_rerun()
    else:
        st.markdown("#### Sign in or create account")
        email = st.text_input("Email")
        password = st.text_input("Password", type="password")
        c1, c2 = st.columns(2)
        with c1:
            if st.button("Sign in"):
                user = verify_user(email, password)
                if user:
                    st.session_state["user"] = user
                    log_event("login", user_id=user["id"])
                    st.experimental_rerun()
                else:
                    st.error("Invalid credentials.")
        with c2:
            if st.button("Create account"):
                if get_user_by_email(email):
                    st.error("Account already exists.")
                else:
                    user = create_user(email, password)
                    st.session_state["user"] = user
                    log_event("signup", user_id=user["id"])
                    st.experimental_rerun()

    if st.session_state["user"]:
        st.markdown("#### Subscription")
        subscribed = bool(st.session_state["user"]["is_subscribed"])
        st.write(f"Status: {'Active' if subscribed else 'Free'}")
        allow_self_subscribe = os.getenv("ALLOW_SELF_SUBSCRIBE", "false").lower() == "true"
        if not subscribed:
            if allow_self_subscribe and st.button("Activate subscription (demo)"):
                set_subscription(st.session_state["user"]["id"], True)
                st.session_state["user"]["is_subscribed"] = 1
                st.success("Subscription activated.")
            else:
                st.caption("Upgrade to unlock Diagnostics, Options, and Export.")

        st.markdown("#### Saved runs")
        runs_df = load_run_summaries(st.session_state["user"]["id"])
        if runs_df.empty:
            st.info("No saved runs yet.")
        else:
            st.dataframe(runs_df, use_container_width=True)

        st.markdown("#### Telemetry (admin)")
        admin_emails = [e.strip().lower() for e in os.getenv("ADMIN_EMAILS", "").split(",") if e.strip()]
        if st.session_state["user"]["email"] in admin_emails:
            conn = get_db_conn()
            events_df = pd.read_sql_query("SELECT * FROM events ORDER BY created_at DESC LIMIT 200", conn)
            conn.close()
            if events_df.empty:
                st.info("No telemetry yet.")
            else:
                events_df["created_at"] = pd.to_datetime(events_df["created_at"], errors="coerce")
                events_df["day"] = events_df["created_at"].dt.date
                daily_active = events_df.groupby("day")["session_id"].nunique().reset_index(name="active_sessions")
                st.markdown("**Daily active sessions**")
                st.line_chart(daily_active.set_index("day")["active_sessions"])

                funnel = events_df["event_name"].value_counts().reindex(
                    ["file_uploaded", "analysis_run", "export_clicked"], fill_value=0
                ).reset_index()
                funnel.columns = ["event", "count"]
                st.markdown("**Upload → analysis → export funnel**")
                st.dataframe(funnel, use_container_width=True)

                error_events = events_df[events_df["event_name"] == "error"]
                if not error_events.empty:
                    st.markdown("**Top errors**")
                    st.dataframe(error_events.head(20), use_container_width=True)
